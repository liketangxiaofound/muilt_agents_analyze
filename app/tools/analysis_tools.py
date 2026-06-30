"""
数据分析工具 — 支持直接读文件或传入 JSON 数据。
大文件时使用 file_path 参数，工具会直接读取完整文件计算，不受预览限制。
"""
import os
import json
import csv
import math
import re
from io import StringIO
from collections import Counter
from strands import tool

from app.tools.storage_tools import process_file_key


# ============================================================
# 文件读取辅助
# ============================================================

def _read_full_file(file_path: str) -> dict:
    """从文件路径读取完整数据，返回统一格式的 dict。
    支持 CSV、Excel、以及 Agent 传入的 JSON 字符串。
    """
    full_path = process_file_key(file_path)
    if not os.path.exists(full_path):
        return {}

    ext = os.path.splitext(full_path)[1].lower()

    if ext == '.csv':
        return _read_csv_full(full_path)
    elif ext in ('.xlsx', '.xls'):
        return _read_excel_full(full_path)
    else:
        return {}


def _read_csv_full(full_path: str) -> dict:
    rows = []
    headers = []
    for enc in ['utf-8-sig', 'utf-8', 'gbk']:
        try:
            with open(full_path, 'r', encoding=enc) as f:
                reader = csv.reader(StringIO(f.read()))
                rows = list(reader)
            break
        except (UnicodeDecodeError, Exception):
            continue
    if not rows:
        return {}
    headers = rows[0]
    data_rows = rows[1:]
    return {"type": "tabular", "source": os.path.basename(full_path),
            "sheets": [{"sheet_name": "data", "row_count": len(data_rows),
                        "column_count": len(headers), "columns": headers, "rows": data_rows}]}


def _read_excel_full(full_path: str) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(full_path, data_only=True)
        sheets = []
        for sname in wb.sheetnames[:1]:  # 只读第一个 sheet
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = [[None if cell is None else str(cell) if isinstance(cell, str) else cell
                          for cell in row] for row in rows[1:]]
            sheets.append({"sheet_name": sname, "row_count": len(data_rows),
                           "column_count": len(headers), "columns": headers, "rows": data_rows})
        wb.close()
        return {"type": "tabular", "source": os.path.basename(full_path), "sheets": sheets}
    except Exception:
        return {}


# ============================================================
# 输入解析 + 列提取
# ============================================================

def _parse_input(data_json):
    if isinstance(data_json, dict):
        return data_json
    if isinstance(data_json, list):
        return data_json[0] if len(data_json) > 0 and isinstance(data_json[0], dict) else {}
    if not isinstance(data_json, str):
        return {}
    try:
        parsed = json.loads(data_json)
        if isinstance(parsed, list):
            return parsed[0] if len(parsed) > 0 and isinstance(parsed[0], dict) else {}
        if isinstance(parsed, dict):
            return parsed
        return {}
    except (json.JSONDecodeError, TypeError):
        if len(data_json.strip()) > 20:
            return {"full_text": data_json, "paragraph_count": 1,
                    "paragraphs": [{"text": data_json, "style": "Normal"}]}
        return {}


def _get_data(data_json: str = "", file_path: str = ""):
    """统一入口：优先用 file_path，其次判断 data_json 是否为文件路径，最后当作 JSON。"""
    if file_path and file_path.strip():
        data = _read_full_file(file_path.strip())
        if data:
            return data
    # data_json 可能就是文件路径（Agent 经常只传第一个参数）
    if data_json and not data_json.strip().startswith("{") and not data_json.strip().startswith("["):
        data = _read_full_file(data_json.strip())
        if data:
            return data
    return _parse_input(data_json)


def _extract_column(data, column):
    values = []
    for sheet in data.get("sheets", []):
        columns = sheet.get("columns", [])
        rows = sheet.get("rows", [])
        if column not in columns:
            continue
        idx = columns.index(column)
        for row in rows:
            if idx < len(row) and row[idx] is not None:
                try:
                    values.append(float(row[idx]))
                except (ValueError, TypeError):
                    pass
    return values


# ============================================================
# 分析工具
# ============================================================

@tool
def statistical_summary(data_json: str = "", file_path: str = "", columns: str = "") -> str:
    """统计摘要。data_json 可以直接传文件路径（如 /mnt/.../file.csv），工具会读取完整文件计算所有行。也支持传入 Fetcher 返回的 JSON 数据。"""
    data = _get_data(data_json, file_path)
    target_cols = [c.strip() for c in columns.split(",") if c.strip()] if columns else None
    all_numeric = set()
    for sheet in data.get("sheets", []):
        for col in sheet.get("columns", []):
            vals = _extract_column(data, col)
            if vals:
                all_numeric.add(col)
    if target_cols is None:
        target_cols = list(all_numeric)
    else:
        target_cols = [c for c in target_cols if c in all_numeric]

    results = {}
    for col in target_cols:
        vals = sorted(_extract_column(data, col))
        if not vals:
            results[col] = {"error": "无有效数值"}
            continue
        n = len(vals)
        mean = sum(vals) / n
        median = vals[n // 2] if n % 2 == 1 else (vals[n // 2 - 1] + vals[n // 2]) / 2
        variance = sum((v - mean) ** 2 for v in vals) / n
        std = math.sqrt(variance)
        results[col] = {"count": n, "mean": round(mean, 2), "median": round(median, 2),
                        "std": round(std, 2), "min": round(vals[0], 2), "max": round(vals[-1], 2),
                        "q1": round(vals[n // 4], 2), "q3": round(vals[3 * n // 4], 2),
                        "range": round(vals[-1] - vals[0], 2)}

    insights = []
    for col, r in results.items():
        if "error" in r:
            continue
        if r["std"] / (abs(r["mean"]) + 0.001) > 0.5:
            insights.append(f"{col} 数据波动较大（标准差={r['std']}），建议关注离群值")
    return json.dumps({"method": "statistical_summary", "result": results, "insights": insights},
                      ensure_ascii=False)


@tool
def anomaly_detection(data_json: str = "", file_path: str = "", column: str = "",
                      threshold_sigma: float = 2.0) -> str:
    """Z-score 异常检测。data_json 可直接传文件路径，工具读取完整文件。"""
    data = _get_data(data_json, file_path)
    all_data = []
    for sheet in data.get("sheets", []):
        columns = sheet.get("columns", [])
        rows = sheet.get("rows", [])
        if column not in columns:
            continue
        idx = columns.index(column)
        for row in rows:
            if idx < len(row) and row[idx] is not None:
                try:
                    all_data.append({"value": float(row[idx]),
                                     "row": {c: row[ci] if ci < len(row) else None
                                             for ci, c in enumerate(columns)}})
                except (ValueError, TypeError):
                    pass
    if not all_data:
        return json.dumps({"error": f"列 '{column}' 无有效数值"}, ensure_ascii=False)

    values = [d["value"] for d in all_data]
    n = len(values)
    mean = sum(values) / n
    std = math.sqrt(sum((v - mean) ** 2 for v in values) / n)
    if std == 0:
        return json.dumps({"method": "anomaly_detection", "column": column, "total_points": n,
                           "anomalies": [], "insights": [f"{column} 所有值均相同（{mean}），无异常"]},
                          ensure_ascii=False)

    anomalies = [{"value": round(d["value"], 2), "z_score": round(abs(d["value"] - mean) / std, 2),
                  "deviation": round(d["value"] - mean, 2), "context": d["row"]}
                 for d in all_data if abs(d["value"] - mean) / std > threshold_sigma]
    return json.dumps({"method": "anomaly_detection", "column": column, "total_points": n,
                       "mean": round(mean, 2), "std": round(std, 2), "threshold_sigma": threshold_sigma,
                       "anomaly_count": len(anomalies), "anomalies": anomalies,
                       "insights": [f"检测到 {len(anomalies)} 个异常值（共 {n} 个数据点）"]},
                      ensure_ascii=False, default=str)


@tool
def trend_analysis(data_json: str = "", file_path: str = "", time_col: str = "",
                   value_col: str = "") -> str:
    """趋势分析：计算变化方向、幅度、拐点。大文件请传 file_path。"""
    data = _get_data(data_json, file_path)
    xs, ys = [], []
    for sheet in data.get("sheets", []):
        columns = sheet.get("columns", [])
        rows = sheet.get("rows", [])
        if time_col not in columns or value_col not in columns:
            continue
        xi = columns.index(time_col)
        yi = columns.index(value_col)
        for row in rows:
            if xi < len(row) and yi < len(row):
                try:
                    ys.append(float(row[yi]))
                    xs.append(row[xi])
                except (ValueError, TypeError):
                    pass
    if len(ys) < 2:
        return json.dumps({"error": "数据点不足"}, ensure_ascii=False)

    n = len(ys)
    x_idx = list(range(n))
    x_mean = (n - 1) / 2
    y_mean = sum(ys) / n
    numerator = sum((x_idx[i] - x_mean) * (ys[i] - y_mean) for i in range(n))
    denominator = sum((x_idx[i] - x_mean) ** 2 for i in range(n))
    slope = numerator / denominator if denominator != 0 else 0

    direction = "上升" if slope > 0.01 * abs(y_mean) else (
        "下降" if slope < -0.01 * abs(y_mean) else "平稳")
    inflection_points = [{"index": i, "time": str(xs[i]), "value": round(ys[i], 2)}
                         for i in range(1, n - 1) if (ys[i] - ys[i - 1]) * (ys[i + 1] - ys[i]) < 0]
    return json.dumps({"method": "trend_analysis", "value_column": value_col,
                       "time_column": time_col, "data_points": n,
                       "first_value": round(ys[0], 2), "last_value": round(ys[-1], 2),
                       "change": round(ys[-1] - ys[0], 2),
                       "change_percent": round((ys[-1] - ys[0]) / (abs(ys[0]) + 0.001) * 100, 1),
                       "direction": direction, "inflection_points": inflection_points,
                       "insights": [f"整体趋势: {direction}（变化 {round((ys[-1] - ys[0]) / (abs(ys[0]) + 0.001) * 100, 1)}%）"]},
                      ensure_ascii=False, default=str)


@tool
def ranking(data_json: str = "", file_path: str = "", column: str = "", top_n: int = 10) -> str:
    """排名分析：Top N 和 Bottom N。大文件请传 file_path。"""
    data = _get_data(data_json, file_path)
    all_rows = []
    for sheet in data.get("sheets", []):
        columns = sheet.get("columns", [])
        rows = sheet.get("rows", [])
        if column not in columns:
            continue
        idx = columns.index(column)
        for row in rows:
            if idx < len(row) and row[idx] is not None:
                try:
                    all_rows.append({"value": float(row[idx]),
                                     "row": {c: row[ci] if ci < len(row) else None
                                             for ci, c in enumerate(columns)}})
                except (ValueError, TypeError):
                    pass
    if not all_rows:
        return json.dumps({"error": f"列 '{column}' 无有效数值"}, ensure_ascii=False)

    sorted_rows = sorted(all_rows, key=lambda x: x["value"], reverse=True)
    n = min(top_n, len(sorted_rows))
    return json.dumps({"method": "ranking", "column": column, "total_items": len(all_rows),
                       "top": [{"rank": i + 1, "value": round(r["value"], 2), "context": r["row"]}
                               for i, r in enumerate(sorted_rows[:n])],
                       "bottom": [{"rank": len(sorted_rows) - i, "value": round(r["value"], 2),
                                   "context": r["row"]}
                                  for i, r in enumerate(reversed(sorted_rows[-n:]))]},
                      ensure_ascii=False, default=str)


@tool
def correlation(data_json: str = "", file_path: str = "", col_a: str = "", col_b: str = "") -> str:
    """两列皮尔逊相关系数。大文件请传 file_path。"""
    data = _get_data(data_json, file_path)
    pairs = []
    for sheet in data.get("sheets", []):
        columns = sheet.get("columns", [])
        if col_a not in columns or col_b not in columns:
            continue
        ai = columns.index(col_a)
        bi = columns.index(col_b)
        for row in sheet.get("rows", []):
            if ai < len(row) and bi < len(row):
                try:
                    pairs.append((float(row[ai]), float(row[bi])))
                except (ValueError, TypeError):
                    pass
    if len(pairs) < 3:
        return json.dumps({"error": "有效数据对不足（需至少 3 对）"}, ensure_ascii=False)

    n = len(pairs)
    mean_a = sum(p[0] for p in pairs) / n
    mean_b = sum(p[1] for p in pairs) / n
    std_a = math.sqrt(sum((p[0] - mean_a) ** 2 for p in pairs) / n)
    std_b = math.sqrt(sum((p[1] - mean_b) ** 2 for p in pairs) / n)
    if std_a == 0 or std_b == 0:
        return json.dumps({"error": "某列方差为 0"}, ensure_ascii=False)

    r = sum((p[0] - mean_a) * (p[1] - mean_b) for p in pairs) / n / (std_a * std_b)
    strength = "强" if abs(r) > 0.7 else ("中等" if abs(r) > 0.4 else ("弱" if abs(r) > 0.2 else "极弱或无"))
    return json.dumps({"method": "correlation", "col_a": col_a, "col_b": col_b,
                       "sample_size": n, "coefficient": round(r, 4), "strength": strength,
                       "direction": "正相关" if r > 0 else "负相关",
                       "insights": [f"{col_a} 与 {col_b} 呈{strength}{'正' if r > 0 else '负'}相关（r={r:.4f}）"]},
                      ensure_ascii=False)


@tool
def text_summary(text_json: str = "", file_path: str = "", max_length: int = 500) -> str:
    """文本关键词和统计。大文件请传 file_path。"""
    data = _get_data(text_json, file_path)
    full_text = data.get("full_text", "")
    if not full_text:
        parts = []
        for p in data.get("paragraphs", []):
            if isinstance(p, str): parts.append(p)
            elif isinstance(p, dict): parts.append(p.get("text", ""))
        full_text = " ".join(parts)
    if not full_text:
        return json.dumps({"error": "无文本内容"}, ensure_ascii=False)

    tokens = [t for t in re.split(r'[，。！？；、\s,.!?;:]+', full_text) if len(t) >= 2]
    word_freq = Counter(tokens).most_common(30)
    return json.dumps({"method": "text_summary", "total_chars": len(full_text),
                       "total_paragraphs": data.get("paragraph_count", 0),
                       "total_tables": data.get("table_count", 0),
                       "truncated": len(full_text) > max_length,
                       "text_preview": full_text[:max_length],
                       "top_keywords": [{"word": w, "count": c} for w, c in word_freq],
                       "insights": [f"文档共 {len(full_text)} 字符，{data.get('paragraph_count', 0)} 段"]},
                      ensure_ascii=False)


@tool
def comparison(data_a_json: str = "", data_b_json: str = "",
               file_a: str = "", file_b: str = "", key_column: str = "") -> str:
    """对比两个数据集。大文件请传 file_a/file_b。"""
    data_a = _get_data(data_a_json, file_a) if (file_a or data_a_json) else {}
    data_b = _get_data(data_b_json, file_b) if (file_b or data_b_json) else {}

    def build_index(d, key):
        idx = {}
        for sheet in d.get("sheets", []):
            columns = sheet.get("columns", [])
            if key not in columns:
                continue
            ki = columns.index(key)
            for row in sheet.get("rows", []):
                if ki < len(row):
                    idx[str(row[ki])] = row
        return idx

    idx_a = build_index(data_a, key_column)
    idx_b = build_index(data_b, key_column)
    keys_a, keys_b = set(idx_a.keys()), set(idx_b.keys())
    common, only_a, only_b = keys_a & keys_b, keys_a - keys_b, keys_b - keys_a
    return json.dumps({"method": "comparison", "key_column": key_column,
                       "total_a": len(keys_a), "total_b": len(keys_b),
                       "common_count": len(common), "only_in_a": list(only_a)[:20],
                       "only_in_b": list(only_b)[:20],
                       "insights": [f"共同项: {len(common)}，仅在 A: {len(only_a)}，仅在 B: {len(only_b)}"]},
                      ensure_ascii=False)


@tool
def classification(text_json: str = "", file_path: str = "", categories: str = "") -> str:
    """对文档段落做预处理分段供后续分类。大文件请传 file_path。"""
    data = _get_data(text_json, file_path)
    paragraphs = data.get("paragraphs", [])
    if not paragraphs:
        return json.dumps({"error": "无段落可分类"}, ensure_ascii=False)

    segments = []
    for i, p in enumerate(paragraphs[:50]):
        if isinstance(p, str):
            text = p.strip(); style = ""
        elif isinstance(p, dict):
            text = p.get("text", "").strip(); style = p.get("style", "")
        else:
            continue
        if len(text) > 20:
            segments.append({"index": i, "text": text[:300], "style": style})

    cat_list = [c.strip() for c in categories.split(",") if c.strip()] if categories else []
    return json.dumps({"method": "classification", "segment_count": len(segments),
                       "provided_categories": cat_list, "segments": segments}, ensure_ascii=False)
